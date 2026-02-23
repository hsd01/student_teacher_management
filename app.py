from flask import Flask, render_template, request, redirect, session, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from db import get_db_connection
from fpdf import FPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from flask import send_file
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "supersecretkey"

def login_required():
    return "user_id" in session

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        print("USER>>>>>",user)
        if user and check_password_hash(user["password"], password):
            if user.get("is_active") == 0:
                return render_template("login.html", error="Your account is suspended. Contact admin.")
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            print(user["id"],":::::::::::::::::::::::::",user["role"])
            if user["role"] == "admin":
                return redirect("/admin")
            return redirect("/dashboard")

        return render_template("login.html", error="Invalid credentials")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# admin login
@app.route("/admin", methods=["GET", "POST"])
def admin_dashboard():
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # create teacher
    if request.method == "POST":
        username = request.form["username"]
        password = generate_password_hash(request.form["password"])
        class_teacher = request.form["class_teacher"]

        cursor.execute("""
            INSERT INTO users (username, password, role, class_teacher, is_active)
            VALUES (%s, %s, 'teacher', %s, 1)
        """, (username, password, class_teacher))

        conn.commit()

    # teachers list
    cursor.execute("SELECT id, username, role, class_teacher, is_active FROM users WHERE role='teacher'")
    teachers = cursor.fetchall()

    # stats
    cursor.execute("SELECT COUNT(*) AS total_students FROM students")
    total_students = cursor.fetchone()["total_students"]

    cursor.execute("SELECT COUNT(*) AS total_teachers FROM users WHERE role='teacher'")
    total_teachers = cursor.fetchone()["total_teachers"]

    cursor.execute("""
        SELECT class, COUNT(*) AS total
        FROM students
        GROUP BY class
        ORDER BY class
    """)
    class_stats = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        "admin_dashboard.html",
        teachers=teachers,
        total_students=total_students,
        total_teachers=total_teachers,
        class_stats=class_stats
    )

# student/stu_id  to view perticular student details 
'''@app.route("/student/<int:id>")
def student_profile(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get teacher class
    cursor.execute("SELECT class_teacher FROM users WHERE id=%s", (session["user_id"],))
    teacher_class = cursor.fetchone()["class_teacher"]

    # Fetch student only if same class
    cursor.execute("SELECT * FROM students WHERE id=%s AND class=%s", (id, teacher_class))
    student = cursor.fetchone()

    cursor.close()
    conn.close()

    if not student:
        return "Access Denied", 403

    return render_template("student_profile.html", student=student)
'''
# add new student(s)
@app.route("/student/add", methods=["POST"])
def add_student():
    if "user_id" not in session:
        return redirect("/")

    name = request.form["name"]
    roll_no = request.form["roll_no"]
    class_name = request.form["class"]

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        "INSERT INTO students (name, roll_no, class) VALUES (%s, %s, %s)",
        (name, roll_no, class_name)
    )

    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/dashboard")

# add all details into student page ui
@app.route("/student/<int:id>")
def student_profile(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Student details
    cur.execute("SELECT * FROM students WHERE id=%s", (id,))
    student = cur.fetchone()

    # Marks per test (total)
    cur.execute("""
        SELECT test_name,
               IFNULL(subject1,0) + IFNULL(subject2,0) + IFNULL(subject3,0) +
               IFNULL(subject4,0) + IFNULL(subject5,0) + IFNULL(subject6,0) AS total
        FROM student_marks
        WHERE student_id=%s
        ORDER BY FIELD(test_name, 'unit1', 'unit2', 'halfyearly', 'final')
    """, (id,))
    rows = cur.fetchall()

    labels = [r["test_name"].upper() for r in rows]
    values = [r["total"] for r in rows]

    cur.close()
    conn.close()
    print("mass iupdate", student)
    return render_template(
        "student_profile.html",
        student=student,
        labels=labels,
        values=values
    )

'''@app.route("/student/<int:id>")
def student_profile(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    cur.execute("SELECT * FROM students WHERE id=%s", (id,))
    student = cur.fetchone()

    cur.execute("""
        SELECT test_name,
               subject1 + subject2 + subject3 + subject4 + subject5 + subject6 AS total
        FROM student_marks WHERE student_id=%s
    """, (id,))
    rows = cur.fetchall()

    labels = [r["test_name"] for r in rows]
    values = [r["total"] for r in rows]

    conn.close()

    return render_template("student_profile.html",
                           student=student,
                           labels=labels,
                           values=values)'''

# import all students
@app.route("/students/import", methods=["POST"])
def import_students():
    if "user_id" not in session:
        return redirect("/")

    file = request.files["file"]
    class_name = request.form["class"]

    df = pd.read_excel(file)
    
    conn = get_db_connection()
    cursor = conn.cursor()

    for _, row in df.iterrows():
        cursor.execute(
            "INSERT INTO students (name, roll_no, class) VALUES (%s, %s, %s)",
            (row["name"], row["roll_no"], class_name)
        )

    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/dashboard")

# student section on gui
'''@app.route("/students")
def students():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    role = session.get("role")
    search = request.args.get("q", "").strip()

    if role == "admin":
        if search:
            cursor.execute("""
                SELECT * FROM students 
                WHERE name LIKE %s
                ORDER BY class, roll_no
            """, (f"%{search}%",))
        else:
            cursor.execute("""
                SELECT * FROM students 
                ORDER BY class, roll_no
            """)
        students = cursor.fetchall()
        teacher_class = None

    else:
        # Teacher – only own class
        cursor.execute("SELECT class_teacher FROM users WHERE id=%s", (session["user_id"],))
        teacher_class = cursor.fetchone()["class_teacher"]

        if search:
            cursor.execute("""
                SELECT * FROM students 
                WHERE class=%s AND name LIKE %s
                ORDER BY roll_no
            """, (teacher_class, f"%{search}%"))
        else:
            cursor.execute("""
                SELECT * FROM students 
                WHERE class=%s 
                ORDER BY roll_no
            """, (teacher_class,))
        students = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        "students.html",
        students=students,
        teacher_class=teacher_class,
        role=role
    )'''
@app.route("/students")
def students_page():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # get role + class
    cur.execute("SELECT role, class_teacher FROM users WHERE id=%s", (session["user_id"],))
    user = cur.fetchone()
    role = user["role"]
    teacher_class = user["class_teacher"]

    search = request.args.get("q", "")
    class_filter = request.args.get("class")

    # For dropdown (admin)
    cur.execute("SELECT DISTINCT class FROM students ORDER BY class")
    classes = cur.fetchall()

    if role == "admin":
        query = "SELECT * FROM students WHERE 1=1"
        params = []

        if search:
            query += " AND name LIKE %s"
            params.append(f"%{search}%")

        if class_filter:
            query += " AND class = %s"
            params.append(class_filter)

        query += " ORDER BY class, roll_no"
        cur.execute(query, params)

    else:
        query = "SELECT * FROM students WHERE class=%s"
        params = [teacher_class]

        if search:
            query += " AND name LIKE %s"
            params.append(f"%{search}%")

        query += " ORDER BY roll_no"
        cur.execute(query, params)

    students = cur.fetchall()
    conn.close()

    return render_template(
        "students.html",
        students=students,
        role=role,
        teacher_class=teacher_class,
        classes=classes
    )

#download student detailtemplate name dob fathers, mothers name etc
@app.route("/students/template_detail")
def download_student_template():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # If admin → export all students
    cursor.execute("SELECT role, class_teacher FROM users WHERE id=%s", (session["user_id"],))
    user = cursor.fetchone()

    if user["role"] == "admin":
        cursor.execute("""
            SELECT name, roll_no, class, email, father_name, mother_name, parent_phone, address, dob
            FROM students
            ORDER BY class, roll_no
        """)
    else:
        # Teacher → only their class
        cursor.execute("""
            SELECT name, class, dob, email, father_name, mother_name, parent_phone, address
            FROM students
            WHERE class = %s
            ORDER BY roll_no
        """, (user["class_teacher"],))

    students = cursor.fetchall()
    cursor.close()
    conn.close()

    # Create Excel
    df = pd.DataFrame(students)

    # Ensure folder exists
    os.makedirs("static/exports", exist_ok=True)

    filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("static/exports", filename)

    df.to_excel(filepath, index=False)

    return send_file(filepath, as_attachment=True)


@app.route("/students/import_detail", methods=["POST"])
def import_students_excel():
    if "user_id" not in session:
        return redirect("/")

    file = request.files["file"]
    df = pd.read_excel(file)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    for _, row in df.iterrows():
        name = str(row["name"]).strip()
        student_class = str(row["class"]).strip()
        dob = row.get("dob")
        email = row.get("email")
        father_name = row.get("father_name")
        mother_name = row.get("mother_name")
        parent_phone = row.get("parent_phone")
        address = row.get("address")

        # 🔍 Find student by (name + class)
        cursor.execute("""
            SELECT id FROM students 
            WHERE name=%s AND class=%s
        """, (name, student_class))
        existing = cursor.fetchone()

        if existing:
            student_id = existing["id"]
            cursor.execute("""
                UPDATE students 
                SET dob=%s, email=%s, father_name=%s, mother_name=%s,
                    parent_phone=%s, address=%s
                WHERE id=%s
            """, (dob, email, father_name, mother_name, parent_phone, address, student_id))
        '''else:
            cursor.execute("""
                INSERT INTO students 
                (name, dob, email, father_name, mother_name, parent_phone, address)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (name, dob, email, father_name, mother_name, parent_phone, address))'''

    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/students")

# bulk move students
@app.route("/admin/students/bulk-move", methods=["POST"])
def bulk_move_students():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Check admin
    cur.execute("SELECT role FROM users WHERE id=%s", (session["user_id"],))
    user = cur.fetchone()
    if user["role"] != "admin":
        conn.close()
        return "Unauthorized", 403

    student_ids = request.form.getlist("student_ids")
    new_class = request.form.get("new_class")

    if not student_ids or not new_class:
        conn.close()
        return redirect("/students")

    placeholders = ",".join(["%s"] * len(student_ids))

    cur.execute(
        f"UPDATE students SET class=%s WHERE id IN ({placeholders})",
        [new_class] + student_ids
    )

    conn.commit()
    conn.close()

    return redirect("/students")


# teachers and admin dashboard
@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # get current user role
    cursor.execute("SELECT role, class_teacher FROM users WHERE id=%s", (session["user_id"],))
    user = cursor.fetchone()

    role = user["role"]
    teacher_class = user["class_teacher"]

    # ==========================
    # 👑 ADMIN DASHBOARD
    # ==========================
    if role == "admin":
        cursor.execute("SELECT id, username, role, class_teacher, is_active FROM users WHERE role='teacher'")
        teachers = cursor.fetchall()

        #total count
        cursor.execute("SELECT COUNT(*) AS total_students FROM students")
        total_students = cursor.fetchone()["total_students"]
        # stats
        cursor.execute("SELECT COUNT(*) AS total_teachers FROM users WHERE role='teacher'")
        total_teachers = cursor.fetchone()["total_teachers"]

        cursor.execute("""
        SELECT class, COUNT(*) AS total
        FROM students
        GROUP BY class
        ORDER BY class
        """)
        class_stats = cursor.fetchall()

        # Students per class
        cursor.execute("""
            SELECT class, COUNT(*) AS total
            FROM students
            GROUP BY class
            ORDER BY class
        """)
        class_stats = cursor.fetchall()

        cursor.close()
        conn.close()
        #cursor.close()
        #conn.close()

        '''return render_template(
            "admin_dashboard.html",
            teachers=teachers,
            total_students=total_students,
            total_teachers=total_teachers,
            class_stats=class_stats
        )'''    
        
        return render_template(
            "admin_dashboard.html",   # 👈 create this template
            total_students=total_students,
            total_teachers=total_teachers,
            class_stats=class_stats,
            teachers=teachers
        )
        
    # ==========================
    # 👨‍🏫 TEACHER DASHBOARD
    # ==========================
    selected_test = request.args.get("test", "unit1")

    # get students of teacher class
    cursor.execute(
        "SELECT id, name, roll_no FROM students WHERE class=%s ORDER BY roll_no ASC",
        (teacher_class,)
    )
    students = cursor.fetchall()

    # get marks for selected test
    cursor.execute("""
        SELECT s.name,
               IFNULL(m.subject1,0)+IFNULL(m.subject2,0)+IFNULL(m.subject3,0)+
               IFNULL(m.subject4,0)+IFNULL(m.subject5,0)+IFNULL(m.subject6,0) AS total
        FROM students s
        LEFT JOIN student_marks m
          ON s.id = m.student_id AND m.test_name = %s
        WHERE s.class = %s
        ORDER BY s.roll_no
    """, (selected_test, teacher_class))

    rows = cursor.fetchall()

    labels = [r["name"] for r in rows]
    values = [r["total"] for r in rows]

    cursor.close()
    conn.close()

    return render_template(
        "teacher_dashboard.html",
        teacher_class=teacher_class,
        students=students,
        labels=labels,
        values=values,
        selected_test=selected_test
    )

# student performance for individual student
@app.route("/performance", methods=["GET", "POST"])
def student_performance():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT class_teacher FROM users WHERE id=%s", (session["user_id"],))
    teacher = cursor.fetchone()
    teacher_class = teacher["class_teacher"]

    cursor.execute("SELECT id, name FROM students WHERE class=%s ORDER BY roll_no ASC", (teacher_class,))
    students = cursor.fetchall()

    selected_student = None
    selected_test = None
    marks = None

    overall_labels = ["Unit 1", "Unit 2", "Half Yearly", "Final"]
    overall_values = [0, 0, 0, 0]

    # ✅ UI-only subject names
    if "subject_names" not in session:
        session["subject_names"] = {}

    if request.method == "POST":

        # Save subject names (UI only)
        if request.form.get("action") == "set_subject_names":
            session["subject_names"] = {
                "sub1": request.form.get("sub1"),
                "sub2": request.form.get("sub2"),
                "sub3": request.form.get("sub3"),
                "sub4": request.form.get("sub4"),
                "sub5": request.form.get("sub5"),
                "sub6": request.form.get("sub6"),
            }

        # View marks
        if request.form.get("action") == "view_marks":
            student_id = request.form["student_id"]
            selected_test = request.form["test_name"]

            cursor.execute("SELECT * FROM students WHERE id=%s", (student_id,))
            selected_student = cursor.fetchone()

            cursor.execute("""
                SELECT subject1, subject2, subject3, subject4, subject5, subject6
                FROM student_marks 
                WHERE student_id=%s AND test_name=%s
            """, (student_id, selected_test))
            marks = cursor.fetchone()

            def get_avg(test_name):
                cursor.execute("""
                    SELECT 
                      COALESCE(AVG(subject1 + subject2 + subject3 + subject4 + subject5 + subject6) / 6, 0) AS avg_marks
                    FROM student_marks
                    WHERE student_id=%s AND test_name=%s
                """, (student_id, test_name))
                row = cursor.fetchone()
                return float(row["avg_marks"]) if row and row["avg_marks"] is not None else 0

            overall_values = [
                get_avg("unit1"),
                get_avg("unit2"),
                get_avg("halfyearly"),
                get_avg("final")
            ]

    cursor.close()
    conn.close()
    total_marks = 0
    average_marks = 0
    percentage = 0

    if marks:
        total_marks = (
            (marks["subject1"] or 0) +
            (marks["subject2"] or 0) +
            (marks["subject3"] or 0) +
            (marks["subject4"] or 0) +
            (marks["subject5"] or 0) +
            (marks["subject6"] or 0)
        )

        average_marks = round(total_marks / 5, 2)

        # assuming each subject is out of 100
        percentage = round((total_marks / 500) * 100, 2)

    return render_template(
        "student_performance.html",
        students=students,
        teacher_class=teacher_class,
        selected_student=selected_student,
        selected_test=selected_test,
        marks=marks,
        overall_labels=overall_labels,
        overall_values=overall_values,
        subject_names=session.get("subject_names", {}),
        total_marks=total_marks,
        average_marks=average_marks,
        percentage=percentage
    )

    '''return render_template(
        "student_performance.html",
        students=students,
        teacher_class=teacher_class,
        selected_student=selected_student,
        selected_test=selected_test,
        marks=marks,
        overall_labels=overall_labels,
        overall_values=overall_values,
        subject_names=session.get("subject_names", {})
    )'''

# edit student(s)
@app.route("/student/edit/<int:id>", methods=["GET", "POST"])
def edit_student(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch existing student
    cursor.execute("SELECT * FROM students WHERE id=%s", (id,))
    student = cursor.fetchone()

    if not student:
        cursor.close()
        conn.close()
        return "Student not found", 404

    if request.method == "POST":
        name = request.form["name"]
        roll_no = request.form["roll_no"]
        student_class = request.form["class"]

        cursor.execute("""
            UPDATE students 
            SET name=%s, roll_no=%s, class=%s
            WHERE id=%s
        """, (name, roll_no, student_class, id))

        conn.commit()
        cursor.close()
        conn.close()

        return redirect("/students")

    cursor.close()
    conn.close()

    return render_template("edit_student.html", student=student)

# delete student 
@app.route("/student/delete/<int:id>")
def delete_student(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM students WHERE id=%s", (id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/students")

# add marks of students for test type like ut1 halfyearly etc
@app.route("/marks/save", methods=["POST"])
def save_marks():
    if "user_id" not in session:
        return redirect("/")

    student_id = request.form["student_id"]
    test_name = request.form["test_name"]
    s1 = request.form["subject1"]
    s2 = request.form["subject2"]
    s3 = request.form["subject3"]
    s4 = request.form["subject4"]
    s5 = request.form["subject5"]
    s6 = request.form["subject6"]

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id FROM student_marks 
        WHERE student_id=%s AND test_name=%s
    """, (student_id, test_name))

    existing = cursor.fetchone()

    if existing:
        cursor.execute("""
            UPDATE student_marks 
            SET subject1=%s, subject2=%s, subject3=%s, subject4=%s, subject5=%s, subject6=%s
            WHERE student_id=%s AND test_name=%s
        """, (s1, s2, s3, s4, s5, s6, student_id, test_name))
    else:
        cursor.execute("""
            INSERT INTO student_marks 
            (student_id, test_name, subject1, subject2, subject3, subject4, subject5, subject6)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (student_id, test_name, s1, s2, s3, s4, s5, s6))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/dashboard")

# import marks from template or excel file
@app.route("/marks/import", methods=["POST"])
def import_marks_excel():
    if "user_id" not in session:
        return redirect("/")

    file = request.files["file"]
    test_name = request.form["test_name"]
    class_name = request.form["class"]

    wb = load_workbook(file)
    sheet = wb.active

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, s1, s2, s3, s4, s5 , s6 = row

        # get student id
        cursor.execute(
            "SELECT id FROM students WHERE name=%s AND class=%s",
            (name, class_name)
        )
        student = cursor.fetchone()

        if not student:
            continue  # skip unknown students

        student_id = student["id"]

        # upsert marks
        cursor.execute("""
            INSERT INTO student_marks (student_id, test_name, subject1, subject2, subject3, subject4, subject5, subject6)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
              subject1=VALUES(subject1),
              subject2=VALUES(subject2),
              subject3=VALUES(subject3),
              subject4=VALUES(subject4),
              subject5=VALUES(subject5),
              subject6=VALUES(subject6)
        """, (student_id, test_name, s1, s2, s3, s4, s5, s6))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/dashboard")

# download marks templates for uploading marks 
@app.route("/marks/template")
def download_marks_template():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get teacher class
    cursor.execute("SELECT class_teacher FROM users WHERE id=%s", (session["user_id"],))
    teacher = cursor.fetchone()
    teacher_class = teacher["class_teacher"]

    # Get students of this class
    cursor.execute("SELECT name, roll_no FROM students WHERE class=%s ORDER BY roll_no", (teacher_class,))
    students = cursor.fetchall()

    cursor.close()
    conn.close()

    # Create Excel file
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"{teacher_class} Marks Template"

    # Header row
    sheet.append(["student_name", "subject1", "subject2", "subject3", "subject4", "subject5", "subject6"])

    # Pre-fill student names
    for s in students:
        sheet.append([s["name"], "", "", "", "", "", ""])

    file_path = f"marks_template_{teacher_class}.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

# download student templates for uploading students name and roll no. in bulk
@app.route("/students/template")
def download_students_template():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Students Template"

    # Header format for students import
    sheet.append(["name", "roll_no"])

    file_path = "students_template.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

# toggel
# ✅ Toggle teacher active / suspended
@app.route("/admin/teacher/toggle/<int:teacher_id>", methods=["POST"])
def toggle_teacher(teacher_id):
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET is_active = IF(is_active = 1, 0, 1) WHERE id = %s",
        (teacher_id,)
    )
    conn.commit()
    cursor.close()
    conn.close()

    return redirect("/admin")


# ❌ Delete teacher
@app.route("/admin/teacher/delete/<int:teacher_id>", methods=["POST"])
def delete_teacher(teacher_id):
    if "user_id" not in session or session.get("role") != "admin":
        return redirect("/")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM users WHERE id = %s", (teacher_id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/admin")

# pdf download 
@app.route("/student/pdf/<int:id>")
def student_pdf(id):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Student details
    cur.execute("SELECT * FROM students WHERE id=%s", (id,))
    student = cur.fetchone()

    # Marks summary
    cur.execute("""
        SELECT test_name,
               IFNULL(subject1,0) + IFNULL(subject2,0) + IFNULL(subject3,0) + 
               IFNULL(subject4,0) + IFNULL(subject5,0) + IFNULL(subject6,0) AS total
        FROM student_marks 
        WHERE student_id=%s
    """, (id,))
    marks = cur.fetchall()

    cur.close()
    conn.close()

    if not student:
        return "Student not found", 404

    pdf = FPDF()
    pdf.add_page()

    # Title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Student Report Card", ln=True, align="C")

    pdf.ln(5)

    # Student Info
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 8, f"Name: {student['name']}", ln=True)
    pdf.cell(0, 8, f"Roll No: {student['roll_no']}", ln=True)
    pdf.cell(0, 8, f"Class: {student['class']}", ln=True)
    pdf.cell(0, 8, f"DOB: {student.get('dob') or 'N/A'}", ln=True)
    pdf.cell(0, 8, f"Father Name: {student.get('father_name') or 'N/A'}", ln=True)
    pdf.cell(0, 8, f"Mother Name: {student.get('mother_name') or 'N/A'}", ln=True)
    pdf.cell(0, 8, f"Parent Contact: {student.get('parent_phone') or 'N/A'}", ln=True)
    pdf.cell(0, 8, f"Email: {student.get('email') or 'N/A'}", ln=True)
    pdf.multi_cell(0, 8, f"Address: {student.get('address') or 'N/A'}")

    pdf.ln(5)

    # Marks Table Header
    pdf.set_font("Arial", "B", 12)
    pdf.cell(60, 8, "Test Name", border=1)
    pdf.cell(40, 8, "Total Marks", border=1, ln=True)

    pdf.set_font("Arial", size=12)

    for m in marks:
        pdf.cell(60, 8, m["test_name"].capitalize(), border=1)
        pdf.cell(40, 8, str(m["total"]), border=1, ln=True)
    total_marks = sum([m["total"] for m in marks]) if marks else 0
    max_marks = len(marks) * 500
    percentage = round((total_marks / max_marks) * 100, 2) if max_marks > 0 else 0

    pdf.ln(5)
    pdf.cell(0, 8, f"Overall Percentage: {percentage}%", ln=True)

    file_path = f"{student['name']}_report_{id}.pdf"
    pdf.output(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/student/pdf/<int:id>/<test_name>")
def student_pdf_term(id, test_name):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Student details
    cur.execute("SELECT * FROM students WHERE id=%s", (id,))
    student = cur.fetchone()

    # Marks for selected term
    cur.execute("""
        SELECT subject1, subject2, subject3, subject4, subject5, subject6
        FROM student_marks 
        WHERE student_id=%s AND test_name=%s
    """, (id, test_name))
    marks = cur.fetchone()

    cur.close()
    conn.close()

    if not student or not marks:
        return "Data not found", 404

    total = sum([marks["subject1"], marks["subject2"], marks["subject3"],
                 marks["subject4"], marks["subject5"], marks["subject6"]])
    percentage = round((total / 500) * 100, 2)

    pdf = FPDF()
    pdf.add_page()

    # Logo
    pdf.image("static/logo.png", x=10, y=8, w=25)

    # Title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "School Report Card", ln=True, align="C")
    pdf.ln(15)

    pdf.set_font("Arial", size=12)
    pdf.cell(0, 8, f"Name: {student['name']}", ln=True)
    pdf.cell(0, 8, f"Class: {student['class']}    Roll No: {student['roll_no']}", ln=True)
    pdf.cell(0, 8, f"Exam: {test_name.capitalize()}", ln=True)

    pdf.ln(5)

    # Marks Table
    pdf.set_font("Arial", "B", 12)
    pdf.cell(50, 8, "Subject", 1)
    pdf.cell(40, 8, "Marks", 1, ln=True)

    pdf.set_font("Arial", size=12)
    for i, sub in enumerate(["Subject 1", "Subject 2", "Subject 3", "Subject 4", "Subject 5", "Subject 6"], start=1):
        pdf.cell(50, 8, sub, 1)
        pdf.cell(40, 8, str(marks[f"subject{i}"]), 1, ln=True)

    pdf.ln(5)
    pdf.cell(0, 8, f"Total: {total} / 500", ln=True)
    pdf.cell(0, 8, f"Percentage: {percentage}%", ln=True)

    pdf.ln(20)
    pdf.cell(0, 8, "Class Teacher Signature: ____________________________", ln=True)

    file_path = f"report_{student['name']}_{test_name}.pdf"
    pdf.output(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/class/pdf/<test_name>")
def class_pdf_bulk(test_name):
    if "user_id" not in session:
        return redirect("/")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # Get teacher class
    cur.execute("SELECT class_teacher FROM users WHERE id=%s", (session["user_id"],))
    teacher = cur.fetchone()
    teacher_class = teacher["class_teacher"]

    # Students
    cur.execute("SELECT * FROM students WHERE class=%s ORDER BY roll_no", (teacher_class,))
    students = cur.fetchall()

    pdf = FPDF()

    for s in students:
        cur.execute("""
            SELECT subject1, subject2, subject3, subject4, subject5, subject6
            FROM student_marks 
            WHERE student_id=%s AND test_name=%s
        """, (s["id"], test_name))
        marks = cur.fetchone()

        if not marks:
            continue

        total = sum(marks.values())
        percentage = round((total / 500) * 100, 2)

        pdf.add_page()
        pdf.image("static/logo.png", x=10, y=8, w=25)

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "School Report Card", ln=True, align="C")
        pdf.ln(15)

        pdf.set_font("Arial", size=12)
        pdf.cell(0, 8, f"Name: {s['name']}", ln=True)
        pdf.cell(0, 8, f"Class: {s['class']}  Roll No: {s['roll_no']}", ln=True)
        pdf.cell(0, 8, f"Exam: {test_name.capitalize()}", ln=True)
        pdf.ln(5)

        for i in range(1, 6):
            pdf.cell(50, 8, f"Subject {i}", 1)
            pdf.cell(40, 8, str(marks[f"subject{i}"]), 1, ln=True)

        pdf.ln(5)
        pdf.cell(0, 8, f"Total: {total} / 500", ln=True)
        pdf.cell(0, 8, f"Percentage: {percentage}%", ln=True)

        pdf.ln(15)
        pdf.cell(0, 8, "Class Teacher Signature: ____________________________", ln=True)

    cur.close()
    conn.close()

    file_path = f"class_{teacher_class}_{test_name}_reports.pdf"
    pdf.output(file_path)

    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001)
